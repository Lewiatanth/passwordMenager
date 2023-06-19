
#Program generujący hasło

import random
import string
import os
from openpyxl import Workbook, load_workbook


password_length = 0
password_minimum_lenght = True
final_password = ""
lowercase_quantity=0
uppercase_quantity=0
special_quantity=0
digits_quantity=0
characters_left = 0
password = []

#Funkcja sprawdzająca istnienie pliku i tworzaca go w razie potrzeby.
def check_file():
    if not os.path.exists("passwords.xlsx"):
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "Witryna"
        ws["B1"] = "Login"
        ws["C1"] = "Hasło"

        wb.save("passwords.xlsx")

#Funkcja, która podczas pierwszego uruchomienia prosi o ustawienie hasła do aplikacji.
def set_app_password(correct_passwords):
    wb = load_workbook("passwords.xlsx")
    ws = wb.active
    if ws.cell(row = 1, column = 4).value is None:
        print("Podczas pierwszego uruchomienia ustaw hasło służące do odczytywania zapisanych danych.")
        while  correct_passwords is False:
            app_password_one = input("Wpisz hasło: ")
            app_password_two = input("Powtórz hasło: ")
            if app_password_one == app_password_two:
                ws["D1"] = app_password_one
                wb.save("passwords.xlsx")
                print("Ustawiono hasło!")
                correct_passwords = True
            else:
                print("Podane hasła są różne!")

#Funkcja wyboru dalszego działania programu.
def choice(choice_status):
    while choice_status is False:
        print("Wybierz co chcesz zrobić:")
        print("1. Wygeneruj nowe hasło.")
        print("2. Odczytaj zapisane hasło.")
        answer = input()
        if answer == "1":
            choice_status = True
            new_password()
        elif answer == "2":
            choice_status = True
            check_app_password()
        else:
            print("Proszę wybrać opcję wpisując 1 lub 2.")

#Funkcja prosząca o podanie długości hasłą i sprawdzająca czy nie jest ona mniejsza od 8.
def enter_password_length(password_length_type):
    global password_length
    global characters_left
    while password_length_type:
        try:
            password = int(input("Podaj długość hasła: "))
        except:
            print("Podany znak nie jest cyfrą.")
        else:
                if password_length < 8:
                    print("Dla bezpieczeństwa hasło nie może składać się z mniej niż 8 znaków.")
                else:
                    characters_left = password_length
                    password_length_type = True

#Funkcja pobierająca ilość: małych i wielkich liter, znaków specjalnych oraz cyfr i weryfikująca poprawność wprowadzonych danych.
def requiment_verify(input_message, error_message, characters_left, type_verify,factor):
    while type_verify:
        try:
            quantity = int(input(input_message))
        except:
            print("Wprowadzone dane nie są liczbą.")
        else:
            if quantity <= 0:
                print(error_message)
            elif quantity > characters_left - factor:
                 print("Wprowadzono ilość większą niż pozostało miejsca w haśle.")
            else:
                characters_left -= quantity
                type_verify = False
    return(characters_left,quantity)

#Funkcja uzupełniająca ilość małych liter o liczbę pozostałych wolnych znaków w haśle.
def check_free_characters():
    global lowercase_quantity
    if characters_left > 0:
         print("Nie wykorzystano wszystkich znaków. Zostaną uzupełnione małymi literami.")
         lowercase_quantity += characters_left

#Funkcja wyświetlająca ustalone przez użytkownika wymagania dotyczące hasła.
def show_conditions():
    print("Długość hasła: ", password_length)
    print("Małe litery: ", lowercase_quantity)
    print("Wielkie litery: ", uppercase_quantity)
    print("Znaki specjalne: ", special_quantity)
    print("Cyfry: ", digits_quantity)

#Kod generujący znaki w haśle.
def generate_password():
    global lowercase_quantity
    global uppercase_quantity
    global special_quantity
    global digits_quantity
    global final_password
    for i in range(password_length):
        if lowercase_quantity > 0:
            password.append(random.choice(string.ascii_lowercase))
            lowercase_quantity -= 1
        if uppercase_quantity > 0:
            password.append(random.choice(string.ascii_uppercase))
            uppercase_quantity -= 1
        if special_quantity > 0:
            password.append(random.choice(string.punctuation))
            special_quantity -= 1
        if digits_quantity > 0:
            password.append(random.choice(string.digits))
            digits_quantity -= 1
        random.shuffle(password)
        final_password = "".join(password)

#Funkcja sprawdzajaca numer pierwszego wolnego wiersza w pierwszej kolumnie.
def check_for_empty_cell():
    wb = load_workbook("passwords.xlsx")
    ws = wb.active
    row = 1
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    return row

#Funkcja zapisująca wygenerowane hasło.
def save_question():
    answer = input("Czy chcesz zapisać hasło? [y/n]")
    if answer == "y":
        row = check_for_empty_cell()
        site = input("Podaj adres strony: ")
        login = input("Podaj login: ")
        wb = load_workbook("passwords.xlsx")
        ws = wb.active
        row_string = str(row)
        site_row = "A" + row_string
        login_row = "B" + row_string
        password_row = "C" + row_string
        ws[site_row] = site
        ws[login_row] = login
        ws[password_row] = final_password
        wb.save("passwords.xlsx")

#Funkcja pozwalająca na stworzenie i zapisanie nowego hasła.
def new_password():
    global characters_left
    global lowercase_quantity
    global uppercase_quantity
    global special_quantity
    global digits_quantity

    enter_password_length(True)
    characters_left, lowercase_quantity = requiment_verify("Podaj ilość małych liter: ", "Hasło musi zawierać minimum 1 małą literę!",characters_left, True, 3)
    characters_left, uppercase_quantity = requiment_verify("Podaj ilość wielkich liter: ", "Hasło musi zawierać minimum 1 wielką literę!", characters_left, True, 2)
    characters_left, special_quantity = requiment_verify("Podaj ilość znaków specjalnych: ","Hasło musi zawierać minimum 1 znak specjalny!",characters_left, True, 1)
    characters_left, digits_quantity = requiment_verify("Podaj ilość cyfr: ", "Hasło musi zawierać minimum 1 cyfrę!",characters_left, True, 0)
    check_free_characters()
    show_conditions()
    generate_password()
    print("Twoje hasło to: ", "".join(password))
    save_question()

#Funkcja pozwalająca na odczytanie istniejącego hasła.
def read_password():
    wb = load_workbook("passwords.xlsx")
    ws = wb.active
    last_row = check_for_empty_cell()
    for i in range(1,last_row-1):
        site = ws.cell(row = i+1, column = 1)
        print(i,site.value)
    answer = int(input("Wpisując cyfrę wybierz witrynę, dla której chcesz odczytać dane logowania: "))
    login = ws.cell(row = answer + 1, column= 2 )
    password = ws.cell(row = answer + 1, column = 3)
    print("Dane logowania:")
    print("Login:",login.value)
    print("Hasło:",password.value)

#Funkcja prosząca o podanie hasła do aplikacji. Jeżeli hasło się nie zgadza to program nie wyświetli danych logowania.
def check_app_password():
    wb = load_workbook("passwords.xlsx")
    ws = wb.active
    answer = input("Aby odczytać dane wpisz hasło: ")
    if answer == ws.cell(row=1, column=4).value:
        read_password()
    else:
        print("Niepoprawne hasło!")
#Wywołanie programu
check_file()
print("Witaj w generatorze haseł!")
set_app_password(False)
choice(False)





